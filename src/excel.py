#!/usr/bin/env python
# -*- coding: UTF-8 -*-

from datetime import datetime, timezone
from enum import IntEnum
from typing import Any, Dict, Iterable, List, Union
import openpyxl
import openpyxl.utils.dataframe
import pandas

import pss_tournament as tourney
import utility as util





# ---------- Constants ----------

__BASE_TABLE_STYLE = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
__EARLIEST_EXCEL_DATETIME = datetime(1900, 1, 1, tzinfo=timezone.utc)


class FILE_ENDING(IntEnum):
    CSV = 1
    JSON = 2
    XL = 3
    XML = 4


__FILE_ENDING_LOOKUP: Dict[FILE_ENDING, str] = {
    FILE_ENDING.CSV: 'csv',
    FILE_ENDING.JSON: 'json',
    FILE_ENDING.XL: 'xlsx',
    FILE_ENDING.XML: 'xml'
}









def create_xl_from_data(data: List[Iterable[Any]], file_prefix: str, data_retrieved_at: datetime, column_formats: List[str], file_name: str = None) -> str:
    if data_retrieved_at is None:
        data_retrieved_at = util.get_utc_now()
    if file_name:
        save_to = file_name
    else:
        save_to = get_file_name(file_prefix, data_retrieved_at, FILE_ENDING.XL)

    wb = openpyxl.Workbook()
    ws = wb.active

    for item in data:
        ws.append(item)

    col_count = len(data[0]) + 1
    row_count = len(data) + 1
    for i, col_no in enumerate(range(1, col_count)):
        column_format = column_formats[i]
        if column_format:
            for row_no in range(2, row_count):
                ws.cell(row_no, col_no).number_format = column_format

    wb.save(save_to)
    return save_to


def create_csv_from_data(data: List[Iterable[Any]], file_prefix: str, data_retrieved_at: datetime, file_name: str = None, delimiter: str = '\t') -> str:
    if data_retrieved_at is None:
        data_retrieved_at = util.get_utc_now()
    if file_name:
        save_to = file_name
    else:
        save_to = get_file_name(file_prefix, data_retrieved_at, FILE_ENDING.CSV)

    if not delimiter:
        delimiter = '\t'

    lines = [
        delimiter.join([
            str(field)
            for field
            in line
        ])
        for line
        in data
    ]

    with open(save_to, mode='w') as fp:
        fp.write('\n'.join(lines))
    return save_to


def create_xl_from_raw_data_dict(flattened_data: List[Iterable[Any]], file_prefix: str, data_retrieved_at: datetime = None) -> str:
    if data_retrieved_at is None:
        data_retrieved_at = util.get_utc_now()
    save_to = get_file_name(file_prefix, data_retrieved_at, FILE_ENDING.XL, consider_tourney=False)

    wb = openpyxl.Workbook(write_only=True)
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.create_sheet()
    df = pandas.DataFrame(flattened_data)
    for (columnName, columnData) in df.iteritems():
        if 'datetime64' in columnData.dtype.name:
            df[columnName] = df[columnName].dt.tz_convert(None)

    for row in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    table = openpyxl.worksheet.table.Table(displayName='tbl', ref=_get_ref_for_df(df))
    table.tableStyleInfo = __BASE_TABLE_STYLE
    table._initialise_columns()
    for cell, col in zip(df.columns, table.tableColumns):
        col.name = str(cell)
    ws.add_table(table)

    wb.save(save_to)
    return save_to


def get_file_name(file_prefix: str, data_retrieved_at: datetime, file_ending: FILE_ENDING, consider_tourney: bool = True) -> str:
    if not file_ending or file_ending not in __FILE_ENDING_LOOKUP.keys():
        file_ending = FILE_ENDING.XL
    file_prefix = file_prefix.replace(' ', '_')
    if consider_tourney and tourney.is_tourney_running(utc_now=data_retrieved_at):
        file_timestamp = f'tournament-{data_retrieved_at.year}-{util.get_month_short_name(data_retrieved_at).lower()}'
    else:
        file_timestamp = data_retrieved_at.strftime('%Y%m%d-%H%M%S')
    suffix = __FILE_ENDING_LOOKUP[file_ending]
    result = f'{file_prefix}_{file_timestamp}.{suffix}'
    return result


def _convert_to_ref(column_count: int, row_count: int, column_start: int = 0, row_start: int = 0, zero_based: bool = True) -> str:
    if zero_based:
        column_start += 1
        row_start += 1
    start_column_letter = openpyxl.utils.get_column_letter(column_start)
    end_column_letter = openpyxl.utils.get_column_letter(column_start + column_count)
    result = f'{start_column_letter}{row_start}:{end_column_letter}{row_start + row_count}'
    return result


def _get_ref_for_df(df: pandas.DataFrame, column_start: int = 0, row_start: int = 0, zero_based: bool = True) -> str:
    col_count = len(df.columns) - 1
    row_count = len(df.index)
    ref = _convert_to_ref(col_count, row_count, column_start=column_start, row_start=row_start, zero_based=zero_based)
    return ref


def fix_field(field: str) -> Union[datetime, int, float, str]:
    if field:
        try:
            dt = util.parse_pss_datetime(field)
            if dt < __EARLIEST_EXCEL_DATETIME:
                dt = __EARLIEST_EXCEL_DATETIME
            dt.replace(tzinfo=None)
            return dt
        except (TypeError, ValueError):
            if not (len(field) >= 2 and field.startswith('0')):
                try:
                    return int(field)
                except (TypeError, ValueError):
                    try:
                        return float(field)
                    except (TypeError, ValueError):
                        pass
        field_lower = field.lower().strip()
        if field_lower == 'false':
            return False
        elif field_lower == 'true':
            return True

    return field